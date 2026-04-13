"""Generate an HTML report for verified equipment from an exported workbook."""

from __future__ import annotations

import argparse
from collections import Counter
from dataclasses import dataclass
from datetime import date, datetime
from html import escape
from pathlib import Path
from statistics import mean
from typing import Iterable

from openpyxl import load_workbook

import app_config as app_config_module

from app_config import APP_CONFIG


DEFAULT_SHEET_NAME = "Inventory"
DEFAULT_OUTPUT_NAME = APP_CONFIG.html_report_filename
TOP_OLDEST_LIMIT = 25
SUMMARY_TABLE_LIMIT = 12
DETAIL_COLUMNS = [
    ("Age", "age", 74, "text"),
    ("Working", "working", 90, "select"),
    ("Lifecycle", "lifecycle", 90, "select"),
    ("Manufacturer", "manufacturer", 130, "text"),
    ("Model", "model", 120, "text"),
    ("Description", "description", 190, "text"),
    ("Serial Number", "serial", 130, "text"),
    ("Location", "location", 130, "text"),
    ("Assigned To", "assigned_to", 105, "text"),
    ("Cal Status", "cal_status", 110, "select"),
    ("Blue Dot", "blue_dot", 88, "text"),
    ("Notes", "notes", 165, "text"),
]


@dataclass(slots=True)
class ReportData:
    source_path: Path
    rows: list[dict[str, str]]
    generated_at: datetime
    by_location: Counter[str]
    by_lifecycle: Counter[str]
    by_working: Counter[str]
    by_cal_status: Counter[str]


@dataclass(frozen=True, slots=True)
class ReviewDecision:
    label: str
    recommendation: str
    css_class: str
    rank: int


def build_report(source_path: Path, sheet_name: str = DEFAULT_SHEET_NAME) -> ReportData:
    """Read verified equipment rows from an exported workbook."""
    workbook = load_workbook(source_path, data_only=True)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(
            f"Sheet '{sheet_name}' not found in {source_path.name}. "
            f"Available sheets: {', '.join(workbook.sheetnames)}"
        )

    sheet = workbook[sheet_name]
    headers = [normalize_header(cell.value) for cell in sheet[1]]
    row_dicts: list[dict[str, str]] = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        values = {header: stringify(value) for header, value in zip(headers, row)}
        if is_verified(values.get("Verified", "")):
            row_dicts.append(values)

    return ReportData(
        source_path=source_path,
        rows=row_dicts,
        generated_at=datetime.now(),
        by_location=counter_from_rows(row_dicts, "Location", empty_label="Unassigned"),
        by_lifecycle=counter_from_rows(row_dicts, "Lifecycle", empty_label="Unknown"),
        by_working=counter_from_rows(row_dicts, "Working", empty_label="Unknown"),
        by_cal_status=counter_from_rows(row_dicts, "Cal Status", empty_label="Unknown"),
    )


def write_report_html(report: ReportData, output_path: Path) -> Path:
    """Write the HTML report to disk."""
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(render_html(report), encoding="utf-8")
    return output_path


def render_html(report: ReportData) -> str:
    """Render the report as a standalone HTML document."""
    sorted_rows = sort_rows_for_review(report.rows)
    oldest_first_rows = sort_rows_by_age_desc(report.rows)
    ages = [age for age in (age_in_years(row) for row in report.rows) if age is not None]
    average_age = f"{mean(ages):.1f}" if ages else "N/A"
    age_20_plus = sum(1 for age in ages if age >= 20)
    age_30_plus = sum(1 for age in ages if age >= 30)
    immediate_review = sum(1 for row in report.rows if review_decision(row).label == "Immediate")
    oldest_rows = "".join(render_oldest_row(row) for row in oldest_first_rows[:TOP_OLDEST_LIMIT])
    detail_rows = "".join(render_detail_row(row) for row in sorted_rows)

    if not oldest_rows:
        oldest_rows = '<tr><td colspan="8" class="empty-state">No verified equipment found.</td></tr>'
    if not detail_rows:
        detail_rows = '<tr><td colspan="13" class="empty-state">No verified equipment found.</td></tr>'

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>{escape(APP_CONFIG.html_report_title)}</title>
  <style>
    body {{ margin: 0; font-family: "Segoe UI", Tahoma, sans-serif; font-size: 16px; background: #f4f6f8; color: #16212b; }}
    .page {{ width: min(1800px, calc(100% - 16px)); margin: 16px auto 32px; }}
    .panel {{ background: #fff; border: 1px solid #d7dde5; padding: 20px; margin-top: 16px; }}
    .header {{ border-top: 6px solid #1f4e79; }}
    h1 {{ margin: 0 0 10px; color: #173954; font-size: 2.1rem; }}
    h2 {{ margin: 0 0 8px; color: #173954; font-size: 1.32rem; }}
    p {{ margin: 0; line-height: 1.6; }}
    .subtext {{ color: #44515f; margin-top: 8px; }}
    .meta {{ margin-top: 14px; display: grid; grid-template-columns: repeat(2, minmax(0, 1fr)); gap: 8px 20px; }}
    .grid {{ display: grid; grid-template-columns: repeat(12, minmax(0, 1fr)); gap: 16px; margin-top: 16px; }}
    .metric {{ grid-column: span 3; min-height: 110px; }}
    .span-5 {{ grid-column: span 5; }}
    .span-7 {{ grid-column: span 7; }}
    .span-12 {{ grid-column: 1 / -1; }}
    .label {{ color: #44515f; text-transform: uppercase; letter-spacing: 0.06em; font-size: 0.96rem; }}
    .value {{ display: block; margin-top: 10px; color: #173954; font-size: 2rem; font-weight: 700; }}
    .note {{ margin-top: 8px; color: #44515f; font-size: 1rem; }}
    .table-shell {{ overflow: auto; border: 1px solid #d7dde5; margin-top: 12px; }}
    table {{ width: 100%; border-collapse: collapse; }}
    th, td {{ text-align: left; padding: 10px 12px; border-bottom: 1px solid #d7dde5; vertical-align: top; font-size: 0.98rem; }}
    th {{ background: #eef3f8; color: #173954; font-size: 0.93rem; text-transform: uppercase; letter-spacing: 0.05em; }}
    tbody tr:nth-child(even) {{ background: #fafbfd; }}
    .status {{ display: inline-block; padding: 4px 8px; border-radius: 4px; font-size: 0.92rem; font-weight: 600; white-space: nowrap; }}
    .status.immediate {{ background: #fde8e8; color: #9b1c1c; }}
    .status.high {{ background: #fff3db; color: #8a4b08; }}
    .status.planned {{ background: #fff8cc; color: #7c5a00; }}
    .status.monitor {{ background: #edf2f7; color: #3f4d5f; }}
    .status.routine {{ background: #e8f4ec; color: #216e39; }}
    .priority-row.immediate td {{ background: #fff1f1; }}
    .priority-row.high td {{ background: #fff7ea; }}
    .priority-row.planned td {{ background: #fffbe6; }}
    .priority-row.monitor td {{ background: #f3f6f9; }}
    .priority-row.routine td {{ background: #eef7f0; }}
    .toolbar {{ display: flex; flex-wrap: wrap; align-items: center; justify-content: space-between; gap: 12px; }}
    .toolbar > div:first-child {{ display: flex; flex-direction: column; justify-content: center; gap: 4px; }}
    .toolbar input {{ width: 240px; min-width: 240px; border: 1px solid #b8c4d0; padding: 10px 12px; font: inherit; color: #16212b; background: #fff; }}
    .toolbar-buttons {{ display: flex; flex-wrap: nowrap; align-items: center; gap: 8px; margin-left: auto; }}
    .toolbar button {{ border: 1px solid #b8c4d0; background: #fff; color: #173954; padding: 9px 12px; font: inherit; cursor: pointer; white-space: nowrap; }}
    .toolbar button:hover {{ background: #eef3f8; }}
    .selection-count {{ color: #44515f; font-size: 1rem; }}
    .panel-toggle {{ display: none; margin-top: 12px; padding: 14px; background: #f8fbfd; border: 1px solid #d7dde5; }}
    .panel-toggle.is-open {{ display: block; }}
    .filter-grid, .column-grid {{ display: grid; gap: 10px 12px; }}
    .filter-grid {{ grid-template-columns: repeat(4, minmax(0, 1fr)); }}
    .column-grid {{ grid-template-columns: repeat(4, minmax(0, 1fr)); }}
    .filter-grid label, .column-grid label {{ display: flex; flex-direction: column; gap: 6px; font-size: 0.96rem; color: #2f3b47; }}
    .column-grid label {{ flex-direction: row; align-items: center; }}
    .filter-grid input, .filter-grid select {{ border: 1px solid #b8c4d0; padding: 8px 10px; font: inherit; color: #16212b; background: #fff; }}
    #verifiedTable {{ table-layout: fixed; font-size: 0.9rem; min-width: 1500px; }}
    #verifiedTable th, #verifiedTable td {{ padding: 7px 9px; line-height: 1.28; white-space: normal; overflow-wrap: anywhere; }}
    #verifiedTable .status {{ padding: 2px 6px; font-size: 0.82rem; }}
    #verifiedTable th {{ position: relative; user-select: none; }}
    #verifiedTable .select-col, #verifiedTable .select-cell {{ width: 42px; text-align: center; }}
    #verifiedTable .resize-handle {{ position: absolute; top: 0; right: 0; width: 14px; height: 100%; cursor: col-resize; }}
    #verifiedTable .resize-handle::before {{ content: ""; position: absolute; left: 5px; top: 8px; bottom: 8px; width: 2px; background: #9aa7b5; border-radius: 999px; box-shadow: 0 0 0 1px #f7f9fb; }}
    #verifiedTable th:hover .resize-handle::before, #verifiedTable .resize-handle:hover::before {{ background: #1f4e79; }}
    #verifiedTable .sort-button {{ display: flex; align-items: center; justify-content: space-between; gap: 8px; width: 100%; padding: 0 22px 0 0; border: 0; background: transparent; color: inherit; font: inherit; text-align: left; cursor: pointer; }}
    #verifiedTable .sort-button:focus-visible {{ outline: 2px solid #1f4e79; outline-offset: 2px; }}
    #verifiedTable .header-label {{ display: block; }}
    #verifiedTable .sort-indicator {{ color: #5f6c7b; font-size: 0.85rem; }}
    #verifiedTable th[aria-sort="ascending"] .sort-indicator::before {{ content: "\\25B2"; }}
    #verifiedTable th[aria-sort="descending"] .sort-indicator::before {{ content: "\\25BC"; }}
    #verifiedTable th[aria-sort="none"] .sort-indicator::before {{ content: "\\2195"; }}
    #verifiedTable input[type="checkbox"] {{ width: 14px; height: 14px; }}
    .col-hidden {{ display: none; }}
    .empty-state {{ text-align: center; color: #5f6c7b; padding: 24px 12px; }}
    @media (max-width: 1000px) {{ .metric, .span-5, .span-7 {{ grid-column: span 6; }} }}
    @media (max-width: 720px) {{
      .page {{ width: min(100% - 12px, 1800px); margin-top: 8px; }}
      .panel {{ padding: 16px; }}
      .meta, .metric, .span-5, .span-7, .span-12 {{ grid-column: 1 / -1; }}
      .filter-grid, .column-grid {{ grid-template-columns: repeat(2, minmax(0, 1fr)); }}
      .toolbar-buttons {{ flex-wrap: wrap; justify-content: flex-start; margin-left: 0; }}
      .toolbar input {{ width: 100%; min-width: 0; }}
    }}
  </style>
</head>
<body>
  <main class="page">
    <section class="panel header">
      <h1>{escape(APP_CONFIG.html_report_title)}</h1>
    </section>

    <section class="grid">
      <article class="panel metric"><div class="label">Verified Equipment</div><span class="value">{len(report.rows)}</span><div class="note">Total verified units included in this review.</div></article>
      <article class="panel metric"><div class="label">Average Age</div><span class="value">{escape(average_age)}</span><div class="note">Average estimated age in years.</div></article>
      <article class="panel metric"><div class="label">30+ Years Old</div><span class="value">{age_30_plus}</span><div class="note">Units that should be part of near-term replacement planning.</div></article>
      <article class="panel metric"><div class="label">Immediate Review</div><span class="value">{immediate_review}</span><div class="note">Units marked scrapped or not working.</div></article>

      <article class="panel span-5">
        <h2>Age Distribution</h2>
        <p class="subtext">This shows how much of the verified inventory sits in older age bands.</p>
        <div class="table-shell"><table><thead><tr><th>Age Band</th><th>Count</th></tr></thead><tbody>{render_age_distribution_rows(report.rows)}</tbody></table></div>
      </article>

      <article class="panel span-7">
        <h2>Manager Action Summary</h2>
        <p class="subtext">These categories summarize what likely needs replacement, recycle, or closer review.</p>
        <div class="table-shell"><table><thead><tr><th>Action Category</th><th>Count</th><th>Meaning</th></tr></thead><tbody>{render_action_summary_rows(report.rows)}</tbody></table></div>
      </article>

      <article class="panel span-12">
        <h2>Oldest Verified Equipment</h2>
        <div class="table-shell"><table><thead><tr><th>Age</th><th>Working</th><th>Lifecycle</th><th>Manufacturer</th><th>Model</th><th>Description</th><th>Serial Number</th><th>Location</th></tr></thead><tbody>{oldest_rows}</tbody></table></div>
      </article>

      <article class="panel span-5">
        <h2>Verified Equipment by Location</h2>
        <p class="subtext">Useful when replacement decisions depend on where the equipment is being used.</p>
        <div class="table-shell"><table><thead><tr><th>Location</th><th>Count</th></tr></thead><tbody>{render_location_rows(report.by_location)}</tbody></table></div>
      </article>

      <article class="panel span-7">
        <h2>Condition Snapshot</h2>
        <p class="subtext">Quick counts for working and lifecycle status in the verified set.</p>
        <div class="table-shell"><table><thead><tr><th>Measure</th><th>Count</th></tr></thead><tbody>
          <tr><td>Working</td><td>{report.by_working.get("working", 0)}</td></tr>
          <tr><td>Not Working</td><td>{report.by_working.get("not_working", 0)}</td></tr>
          <tr><td>Limited</td><td>{report.by_working.get("limited", 0)}</td></tr>
          <tr><td>Unknown Working Status</td><td>{report.by_working.get("unknown", 0)}</td></tr>
          <tr><td>Active Lifecycle</td><td>{report.by_lifecycle.get("active", 0)}</td></tr>
          <tr><td>Scrapped Lifecycle</td><td>{report.by_lifecycle.get("scrapped", 0)}</td></tr>
          <tr><td>20+ Years Old</td><td>{age_20_plus}</td></tr>
          <tr><td>30+ Years Old</td><td>{age_30_plus}</td></tr>
        </tbody></table></div>
      </article>

      <article class="panel span-12">
        <div class="toolbar">
          <div>
            <h2>Full Verified Equipment Detail</h2>
            <div class="selection-count" id="selectionCount">0 selected</div>
          </div>
          <div class="toolbar-buttons">
            <button type="button" id="toggleFilters">Filters</button>
            <button type="button" id="toggleColumns">Columns</button>
            <button type="button" id="clearSelection">Clear Selection</button>
            <input id="tableFilter" type="search" placeholder="Search">
          </div>
        </div>
        <div class="panel-toggle" id="filtersPanel">
          <div class="filter-grid">
            {render_filter_controls()}
          </div>
        </div>
        <div class="panel-toggle" id="columnsPanel">
          <div class="column-grid">
            {render_column_toggles()}
          </div>
        </div>
        <div class="table-shell"><table id="verifiedTable"><colgroup>{render_colgroup()}</colgroup><thead><tr><th class="select-col"><input type="checkbox" id="selectAllRows"></th>{render_detail_headers()}</tr></thead><tbody>{detail_rows}</tbody></table></div>
      </article>
    </section>
  </main>
  <script>
    const filterInput = document.getElementById("tableFilter");
    const tableRows = Array.from(document.querySelectorAll("#verifiedTable tbody tr"));
    const detailTable = document.getElementById("verifiedTable");
    const detailTableBody = detailTable.querySelector("tbody");
    const filtersPanel = document.getElementById("filtersPanel");
    const columnsPanel = document.getElementById("columnsPanel");
    const selectionCount = document.getElementById("selectionCount");
    const selectAllRows = document.getElementById("selectAllRows");
    const rowCheckboxes = Array.from(document.querySelectorAll(".row-select"));
    const filterControls = Array.from(document.querySelectorAll(".column-filter"));
    const columnToggles = Array.from(document.querySelectorAll(".column-toggle"));
    const sortableHeaders = Array.from(document.querySelectorAll("#verifiedTable th[data-sortable='true']"));
    const collator = new Intl.Collator(undefined, {{ numeric: true, sensitivity: "base" }});
    const numericColumns = new Set(["age", "blue_dot"]);
    let activeSort = {{ columnKey: null, direction: "ascending" }};

    function updateSelectionCount() {{
      const selected = rowCheckboxes.filter((box) => box.checked).length;
      selectionCount.textContent = `${{selected}} selected`;
    }}

    function applyFilters() {{
      const query = filterInput.value.trim().toLowerCase();
      tableRows.forEach((row) => {{
        let visible = row.innerText.toLowerCase().includes(query);
        if (visible) {{
          for (const control of filterControls) {{
            const key = control.dataset.column;
            const value = control.value.trim().toLowerCase();
            if (!value) {{
              continue;
            }}
            const cell = row.querySelector(`[data-column="${{key}}"]`);
            const text = cell ? cell.innerText.trim().toLowerCase() : "";
            const isSelectFilter = control.tagName === "SELECT";
            if ((isSelectFilter && text !== value) || (!isSelectFilter && !text.includes(value))) {{
              visible = false;
              break;
            }}
          }}
        }}
        row.style.display = visible ? "" : "none";
      }});
      const visibleRows = tableRows.filter((row) => row.style.display !== "none");
      const allVisibleSelected = visibleRows.length > 0 && visibleRows.every((row) => row.querySelector(".row-select").checked);
      selectAllRows.checked = allVisibleSelected;
      selectAllRows.indeterminate = visibleRows.length > 0 && !allVisibleSelected && visibleRows.some((row) => row.querySelector(".row-select").checked);
    }}

    function setColumnVisibility(columnKey, visible) {{
      document.querySelectorAll(`.col-${{columnKey}}`).forEach((cell) => {{
        cell.classList.toggle("col-hidden", !visible);
      }});
      document.querySelectorAll(`col[data-column="${{columnKey}}"]`).forEach((col) => {{
        col.style.display = visible ? "" : "none";
      }});
    }}

    function cellText(row, columnKey) {{
      const cell = row.querySelector(`[data-column="${{columnKey}}"]`);
      if (!cell) {{
        return "";
      }}
      const text = cell.textContent.trim();
      return text === "-" ? "" : text;
    }}

    function updateSortIndicators(columnKey, direction) {{
      sortableHeaders.forEach((header) => {{
        const ariaSort = header.dataset.column === columnKey ? direction : "none";
        header.setAttribute("aria-sort", ariaSort);
      }});
    }}

    function sortDetailRows(columnKey, direction) {{
      const sorted = [...tableRows].sort((rowA, rowB) => {{
        const valueA = cellText(rowA, columnKey);
        const valueB = cellText(rowB, columnKey);

        if (numericColumns.has(columnKey)) {{
          const numberA = Number.parseFloat(valueA);
          const numberB = Number.parseFloat(valueB);
          const aIsNumber = Number.isFinite(numberA);
          const bIsNumber = Number.isFinite(numberB);

          if (aIsNumber && bIsNumber && numberA !== numberB) {{
            return direction === "ascending" ? numberA - numberB : numberB - numberA;
          }}
          if (aIsNumber !== bIsNumber) {{
            return aIsNumber ? -1 : 1;
          }}
        }}

        const comparison = collator.compare(valueA, valueB);
        return direction === "ascending" ? comparison : -comparison;
      }});

      detailTableBody.append(...sorted);
      activeSort = {{ columnKey, direction }};
      updateSortIndicators(columnKey, direction);
    }}

    document.getElementById("toggleFilters").addEventListener("click", () => {{
      filtersPanel.classList.toggle("is-open");
    }});

    document.getElementById("toggleColumns").addEventListener("click", () => {{
      columnsPanel.classList.toggle("is-open");
    }});

    document.getElementById("clearSelection").addEventListener("click", () => {{
      rowCheckboxes.forEach((box) => {{
        box.checked = false;
      }});
      selectAllRows.checked = false;
      selectAllRows.indeterminate = false;
      updateSelectionCount();
    }});

    filterInput.addEventListener("input", applyFilters);
    filterControls.forEach((control) => control.addEventListener("input", applyFilters));
    filterControls.forEach((control) => control.addEventListener("change", applyFilters));
    columnToggles.forEach((toggle) => toggle.addEventListener("change", (event) => {{
      setColumnVisibility(event.target.dataset.column, event.target.checked);
    }}));
    sortableHeaders.forEach((header) => {{
      const button = header.querySelector(".sort-button");
      if (!button) {{
        return;
      }}
      button.addEventListener("click", () => {{
        const columnKey = header.dataset.column;
        const direction =
          activeSort.columnKey === columnKey && activeSort.direction === "ascending"
            ? "descending"
            : "ascending";
        sortDetailRows(columnKey, direction);
      }});
    }});

    selectAllRows.addEventListener("change", () => {{
      tableRows.forEach((row) => {{
        if (row.style.display === "none") {{
          return;
        }}
        row.querySelector(".row-select").checked = selectAllRows.checked;
      }});
      updateSelectionCount();
    }});

    rowCheckboxes.forEach((box) => box.addEventListener("change", () => {{
      updateSelectionCount();
      applyFilters();
    }}));

    document.querySelectorAll("#verifiedTable th[data-resizable='true']").forEach((header) => {{
      const handle = header.querySelector(".resize-handle");
      const key = header.dataset.column;
      if (!handle || !key) {{
        return;
      }}
      handle.addEventListener("mousedown", (event) => {{
        event.preventDefault();
        const startX = event.clientX;
        const col = document.querySelector(`col[data-column="${{key}}"]`);
        const startWidth = col ? (parseFloat(col.style.width) || header.offsetWidth) : header.offsetWidth;
        const onMove = (moveEvent) => {{
          const nextWidth = Math.max(60, startWidth + (moveEvent.clientX - startX));
          if (col) {{
            col.style.width = `${{nextWidth}}px`;
          }}
        }};
        const onUp = () => {{
          window.removeEventListener("mousemove", onMove);
          window.removeEventListener("mouseup", onUp);
        }};
        window.addEventListener("mousemove", onMove);
        window.addEventListener("mouseup", onUp);
      }});
    }});

    updateSelectionCount();
    applyFilters();
  </script>
</body>
</html>
"""


def render_detail_headers() -> str:
    headers: list[str] = []
    for label, key, _, _ in DETAIL_COLUMNS:
        headers.append(
            f'<th class="col-{escape(key)}" data-column="{escape(key)}" '
            'data-resizable="true" data-sortable="true" aria-sort="none">'
            '<button type="button" class="sort-button">'
            f'<span class="header-label">{escape(label)}</span>'
            '<span class="sort-indicator" aria-hidden="true"></span>'
            "</button>"
            '<span class="resize-handle"></span>'
            "</th>"
        )
    return "".join(headers)


def render_colgroup() -> str:
    cols = ['<col class="select-col" style="width: 42px;">']
    for _, key, width, _ in DETAIL_COLUMNS:
        cols.append(f'<col class="col-{escape(key)}" data-column="{escape(key)}" style="width: {width}px;">')
    return "".join(cols)


def render_filter_controls() -> str:
    controls: list[str] = []
    select_options = {
        "working": ["", "working", "limited", "not_working", "unknown"],
        "lifecycle": ["", "active", "repair", "scrapped", "missing", "rental"],
        "cal_status": ["", "calibrated", "reference_only", "out_to_cal", "unknown"],
    }
    for label, key, _, filter_type in DETAIL_COLUMNS:
        if filter_type == "select":
            options = "".join(
                f'<option value="{escape(pretty_label(value) if value else "")}">'
                f'{escape(pretty_label(value) if value else "All")}</option>'
                for value in select_options[key]
            )
            controls.append(
                f'<label>{escape(label)}<select class="column-filter" data-column="{escape(key)}">{options}</select></label>'
            )
        else:
            controls.append(
                f'<label>{escape(label)}<input class="column-filter" data-column="{escape(key)}" type="text"></label>'
            )
    return "".join(controls)


def render_column_toggles() -> str:
    toggles: list[str] = []
    for label, key, _, _ in DETAIL_COLUMNS:
        toggles.append(
            f'<label><input class="column-toggle" data-column="{escape(key)}" type="checkbox" checked> {escape(label)}</label>'
        )
    return "".join(toggles)


def render_oldest_row(row: dict[str, str]) -> str:
    decision = review_decision(row)
    return (
        f'<tr class="priority-row {escape(decision.css_class)}">'
        f"<td>{escape(display_or_dash(row.get('Est. Age (Yrs)', '')))}</td>"
        f"<td>{escape(pretty_label(display_or_dash(row.get('Working', ''))))}</td>"
        f"<td>{escape(pretty_label(display_or_dash(row.get('Lifecycle', ''))))}</td>"
        f"<td>{escape(display_or_dash(row.get('Manufacturer', '')))}</td>"
        f"<td>{escape(display_or_dash(row.get('Model', '')))}</td>"
        f"<td>{escape(display_or_dash(row.get('Description', '')))}</td>"
        f"<td>{escape(display_or_dash(row.get('Serial Number', '')))}</td>"
        f"<td>{escape(display_or_dash(row.get('Location', '')))}</td>"
        "</tr>"
    )


def render_detail_row(row: dict[str, str]) -> str:
    decision = review_decision(row)
    field_map = {
        "age": "Est. Age (Yrs)",
        "working": "Working",
        "lifecycle": "Lifecycle",
        "manufacturer": "Manufacturer",
        "model": "Model",
        "description": "Description",
        "serial": "Serial Number",
        "location": "Location",
        "assigned_to": "Assigned To",
        "cal_status": "Cal Status",
        "blue_dot": "Blue Dot",
        "notes": "Notes",
    }
    cells = ['<td class="select-cell"><input class="row-select" type="checkbox"></td>']
    for _, key, _, _ in DETAIL_COLUMNS:
        value = row.get(field_map[key], "")
        value = pretty_label(display_or_dash(value)) if key in {"working", "lifecycle", "cal_status"} else display_or_dash(value)
        cells.append(f'<td class="col-{escape(key)}" data-column="{escape(key)}">{escape(value)}</td>')
    return f'<tr class="priority-row {escape(decision.css_class)}">{"".join(cells)}</tr>'


def render_age_distribution_rows(rows: Iterable[dict[str, str]]) -> str:
    counts = Counter(age_band(row) for row in rows)
    ordered = ["40+ years", "30-39 years", "20-29 years", "10-19 years", "Under 10 years", "Age not available"]
    return "".join(f"<tr><td>{band}</td><td>{counts.get(band, 0)}</td></tr>" for band in ordered)


def render_action_summary_rows(rows: Iterable[dict[str, str]]) -> str:
    counts = Counter(review_decision(row).label for row in rows)
    meanings = {
        "Immediate": "Scrapped or not working. Decide quickly whether to recycle, replace, or repair.",
        "High": "Very old equipment. Replacement should be prioritized soon.",
        "Planned": "Old enough that it should be included in the budget pipeline.",
        "Monitor": "Older or uncertain condition. Review before the next budget cycle.",
        "Routine": "No strong replacement signal from current age and status fields alone.",
    }
    labels = ["Immediate", "High", "Planned", "Monitor", "Routine"]
    return "".join(
        f"<tr><td>{render_status(label, label.lower())}</td><td>{counts.get(label, 0)}</td><td>{escape(meanings[label])}</td></tr>"
        for label in labels
    )


def render_location_rows(counter: Counter[str]) -> str:
    if not counter:
        return '<tr><td colspan="2" class="empty-state">No verified equipment found.</td></tr>'
    return "".join(f"<tr><td>{escape(label)}</td><td>{count}</td></tr>" for label, count in counter.most_common(SUMMARY_TABLE_LIMIT))


def sort_rows_for_review(rows: Iterable[dict[str, str]]) -> list[dict[str, str]]:
    return sorted(
        rows,
        key=lambda row: (
            review_decision(row).rank,
            -(age_in_years(row) or -1),
            row.get("Manufacturer", "").lower(),
            row.get("Model", "").lower(),
            row.get("Serial Number", "").lower(),
        ),
    )


def sort_rows_by_age_desc(rows: Iterable[dict[str, str]]) -> list[dict[str, str]]:
    return sorted(
        rows,
        key=lambda row: (
            age_in_years(row) is None,
            -(age_in_years(row) or 0),
            row.get("Manufacturer", "").lower(),
            row.get("Model", "").lower(),
            row.get("Serial Number", "").lower(),
        ),
    )


def review_decision(row: dict[str, str]) -> ReviewDecision:
    age = age_in_years(row)
    lifecycle = row.get("Lifecycle", "").strip().lower()
    working = row.get("Working", "").strip().lower()

    if lifecycle == "scrapped":
        return ReviewDecision("Immediate", "Marked scrapped. Remove from active use and replace only if still needed.", "immediate", 0)
    if working == "not_working":
        return ReviewDecision("Immediate", "Not working. Compare repair cost against replacement value.", "immediate", 1)
    if age is not None and age >= 40:
        return ReviewDecision("High", "Very old equipment. Prioritize replacement planning.", "high", 2)
    if age is not None and age >= 30:
        return ReviewDecision("Planned", "Old equipment. Include in near-term budget planning.", "planned", 3)
    if working in {"limited", "unknown"} or (age is not None and age >= 20):
        return ReviewDecision("Monitor", "Review condition and future need before the next budget cycle.", "monitor", 4)
    return ReviewDecision("Routine", "No immediate replacement signal from the current export fields.", "routine", 5)


def render_status(label: str, css_class: str) -> str:
    return f'<span class="status {escape(css_class)}">{escape(label)}</span>'


def age_band(row: dict[str, str]) -> str:
    age = age_in_years(row)
    if age is None:
        return "Age not available"
    if age >= 40:
        return "40+ years"
    if age >= 30:
        return "30-39 years"
    if age >= 20:
        return "20-29 years"
    if age >= 10:
        return "10-19 years"
    return "Under 10 years"


def age_in_years(row: dict[str, str]) -> float | None:
    raw_value = row.get("Est. Age (Yrs)", "").strip()
    if not raw_value:
        return None
    try:
        return float(raw_value)
    except ValueError:
        return None


def display_or_dash(value: str) -> str:
    return value.strip() if value and value.strip() else "-"


def counter_from_rows(
    rows: Iterable[dict[str, str]],
    column_name: str,
    *,
    empty_label: str,
) -> Counter[str]:
    counter: Counter[str] = Counter()
    for row in rows:
        label = row.get(column_name, "").strip() or empty_label
        counter[label] += 1
    return counter


def normalize_header(value: object) -> str:
    return str(value).strip() if value is not None else ""


def stringify(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return f"{value:.1f}".rstrip("0").rstrip(".")
    return str(value).strip()


def is_verified(value: object) -> bool:
    normalized = stringify(value).strip().lower()
    return normalized in {"yes", "y", "true", "1", "verified"}

def pretty_label(value: str) -> str:
    return value.replace("_", " ").strip().title()


def app_root() -> Path:
    """Return the active app root so report defaults stay app-specific."""
    config_file = getattr(app_config_module, "__file__", "")
    if config_file:
        return Path(config_file).resolve().parent
    return Path.cwd()


def default_export_dir() -> Path:
    return app_root() / "Output"


def default_output_path(source_path: Path) -> Path:
    return app_root() / "Output" / f"{source_path.stem}_{DEFAULT_OUTPUT_NAME}"


def find_latest_workbook(export_dir: Path) -> Path:
    workbooks = sorted(
        (
            path for path in export_dir.glob("*.xlsx")
            if not path.name.startswith("~$")
        ),
        key=lambda path: path.stat().st_mtime,
        reverse=True,
    )
    if not workbooks:
        raise FileNotFoundError(f"No .xlsx files found in {export_dir}")
    return workbooks[0]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Generate an HTML report for verified equipment from an exported workbook."
    )
    parser.add_argument(
        "--input",
        type=Path,
        help="Path to the exported .xlsx file. Defaults to the newest workbook in the app Output folder.",
    )
    parser.add_argument(
        "--output",
        type=Path,
        help="Output HTML path. Defaults to the app Output folder.",
    )
    parser.add_argument(
        "--sheet",
        default=DEFAULT_SHEET_NAME,
        help=f"Workbook sheet to read. Defaults to {DEFAULT_SHEET_NAME!r}.",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()

    source_path = args.input or find_latest_workbook(default_export_dir())
    output_path = args.output or default_output_path(source_path)

    report = build_report(source_path, sheet_name=args.sheet)
    write_report_html(report, output_path)

    print(f"Verified equipment report written to: {output_path}")
    print(f"Verified equipment rows included: {len(report.rows)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

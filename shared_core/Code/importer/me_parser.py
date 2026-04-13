"""Parser for the ME single-workbook inventory import."""

from __future__ import annotations

from pathlib import Path

import openpyxl

from app_config import APP_CONFIG
from Code.db.models import Equipment, ImportIssue, RawCell
from Code.importer.normalizer import clean_value, col_to_letter, is_placeholder

MASTER_FILE = APP_CONFIG.master_source_file

_HEADER_ALIASES = {
    "item no.": "item_no",
    "item no": "item_no",
    "material name": "material_name",
    "description": "description",
    "purpouse": "purpose",
    "purpose": "purpose",
    "asset tag id": "asset_tag_id",
    "model": "model",
    "model no.": "model",
    "serial number": "serial_number",
    "quantity": "qty",
    "location": "location",
    "picture": "picture",
    "box no.": "box_no",
    "box no": "box_no",
}


def parse_me_workbook(wb_path: Path) -> tuple[list[Equipment], list[ImportIssue]]:
    """Parse the ME workbook into normalized inventory records."""
    wb = openpyxl.load_workbook(str(wb_path), data_only=True)
    records: list[Equipment] = []
    issues: list[ImportIssue] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        header_row, col_map = _find_header_row(ws)
        normalized_sheet = _normalize_sheet_name(sheet_name)

        if header_row < 0:
            issues.append(
                ImportIssue(
                    issue_type="parse_error",
                    source_file=MASTER_FILE,
                    source_sheet=sheet_name,
                    source_row=0,
                    summary=f"Could not find a supported header row in '{sheet_name}'",
                )
            )
            continue

        for row_idx in range(header_row + 1, ws.max_row + 1):
            row = _row_to_dict(ws, row_idx, col_map)
            if _is_empty_row(row):
                continue

            material_name = row.get("material_name", "")
            description = row.get("description", "")
            if not material_name and not description:
                continue

            asset_number = row.get("asset_tag_id", "")
            if is_placeholder(asset_number):
                asset_number = ""

            serial_number = row.get("serial_number", "")
            if is_placeholder(serial_number):
                serial_number = ""

            import_key = _build_import_key(
                normalized_sheet,
                row.get("item_no", ""),
                material_name,
                row.get("model", ""),
                row.get("location", ""),
                asset_number,
                serial_number,
            )
            notes = _build_notes(sheet_name, row)

            eq = Equipment(
                asset_number=asset_number,
                serial_number=serial_number,
                manufacturer=material_name,
                manufacturer_raw=material_name,
                model=row.get("model", ""),
                description=description or material_name,
                qty=_safe_float(row.get("qty", "")),
                location=row.get("location", ""),
                notes=notes,
            )
            eq.add_source_ref(MASTER_FILE, sheet_name, row_idx, import_key=import_key)
            records.append(eq)

    return records, issues


def index_me_raw_cells(wb_path: Path) -> list[RawCell]:
    """Index every non-empty cell in the ME workbook for raw search."""
    wb = openpyxl.load_workbook(str(wb_path), data_only=True)
    cells: list[RawCell] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row_idx in range(1, ws.max_row + 1):
            row_vals = []
            for col_idx in range(1, ws.max_column + 1):
                value = clean_value(ws.cell(row=row_idx, column=col_idx).value)
                if value:
                    row_vals.append(value)

            row_preview = " | ".join(row_vals[:6])
            for col_idx in range(1, ws.max_column + 1):
                value = clean_value(ws.cell(row=row_idx, column=col_idx).value)
                if value and not value.isspace():
                    cells.append(
                        RawCell(
                            source_file=MASTER_FILE,
                            source_sheet=sheet_name,
                            row_number=row_idx,
                            column_number=col_idx,
                            cell_address=f"{col_to_letter(col_idx - 1)}{row_idx}",
                            cell_value=value,
                            row_preview=row_preview,
                        )
                    )

    return cells


def _find_header_row(ws, max_rows: int = 5) -> tuple[int, dict[str, int]]:
    """Find the header row and return a canonical field map."""
    for row_idx in range(1, min(max_rows, ws.max_row) + 1):
        row_vals = [clean_value(cell.value).lower() for cell in ws[row_idx]]
        col_map: dict[str, int] = {}
        for col_idx, value in enumerate(row_vals, start=1):
            canonical = _HEADER_ALIASES.get(value)
            if canonical and canonical not in col_map:
                col_map[canonical] = col_idx
        if "material_name" in col_map and "qty" in col_map:
            return row_idx, col_map
    return -1, {}


def _row_to_dict(ws, row_idx: int, col_map: dict[str, int]) -> dict[str, str]:
    """Extract a single sheet row as a canonical field dictionary."""
    result: dict[str, str] = {}
    for field, col_idx in col_map.items():
        result[field] = clean_value(ws.cell(row=row_idx, column=col_idx).value)
    return result


def _is_empty_row(row: dict[str, str]) -> bool:
    """Return whether the parsed row contains only blank or placeholder values."""
    return all(not value or is_placeholder(value) for value in row.values())


def _build_notes(sheet_name: str, row: dict[str, str]) -> str:
    """Flatten ME-only fields into notes until the ME schema is expanded."""
    parts = [f"Source Sheet: {sheet_name}"]

    description = row.get("description", "")
    if description:
        parts.append(f"Source Description: {description}")

    purpose = row.get("purpose", "")
    if purpose:
        parts.append(f"Purpose: {purpose}")

    box_no = row.get("box_no", "")
    if box_no and not is_placeholder(box_no):
        parts.append(f"Box No: {box_no}")

    picture = row.get("picture", "")
    if picture and not is_placeholder(picture):
        parts.append(f"Picture: {picture}")

    return " | ".join(parts)


def _normalize_sheet_name(sheet_name: str) -> str:
    """Create a stable sheet slug for ME import keys."""
    return "_".join(sheet_name.strip().lower().split())


def _build_import_key(
    sheet_slug: str,
    item_no: str,
    material_name: str,
    model: str,
    location: str,
    asset_number: str,
    serial_number: str,
) -> str:
    """Build a stable ME import key that survives quantity changes."""
    parts = [
        sheet_slug,
        _slug(item_no),
        _slug(material_name),
        _slug(model),
        _slug(location),
        _slug(asset_number),
        _slug(serial_number),
    ]
    meaningful = [part for part in parts if part]
    return "me::" + "::".join(meaningful)


def _slug(value: str) -> str:
    """Normalize a workbook field for stable import-key use."""
    return "-".join(clean_value(value).lower().split())


def _safe_float(value: str) -> float | None:
    """Convert a workbook value to float when possible."""
    if not value or is_placeholder(value):
        return None
    try:
        return float(value)
    except ValueError:
        return None

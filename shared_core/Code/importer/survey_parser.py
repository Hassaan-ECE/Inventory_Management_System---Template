"""Parser for the Survey workbook (.xlsx)."""

import json
from pathlib import Path

import openpyxl

from app_config import APP_CONFIG
from Code.db.models import ImportIssue, RawCell
from Code.importer.normalizer import (
    clean_value, col_to_letter, is_placeholder, normalize_manufacturer,
)

SURVEY_FILE = APP_CONFIG.survey_source_file

# Section labels in the survey (lowercase for matching)
SECTION_LABELS = {
    "may 2025 calibrated":     "may_2025_cal",
    "oct 2025 cal":            "oct_2025_cal",
    "non cal equip":           "non_cal",
    "eng automated test sets": "ats_calibrated",
    "non working equipment":   "non_working",
}

# Expected column layout (0-based from the data I inspected)
# Col A (1): Blue Dot
# Col B (2): Asset Numbers
# Col C (3): Serial Number
# Col D (4): Manufacturer
# Col E (5): Model
# Col F (6): Description
# Col G (7): Chk List
# Col H (8): Location
# Col I (9): How Old is Equip (Years)
SURVEY_COLS = {
    "blue_dot":       1,
    "asset_number":   2,
    "serial_number":  3,
    "manufacturer":   4,
    "model":          5,
    "description":    6,
    "chk_list":       7,
    "location":       8,
    "how_old":        9,
}


def _detect_section(cell_values: list[str]) -> str | None:
    """If this row is a section label, return the section key. Otherwise None."""
    joined = " ".join(v.strip().lower() for v in cell_values if v)
    for label, key in SECTION_LABELS.items():
        if label in joined:
            return key
    return None


def _is_header_row(cell_values: list[str]) -> bool:
    """Check if this looks like the column header row."""
    joined = " ".join(v.strip().lower() for v in cell_values if v)
    return "asset number" in joined or "blue dot" in joined


def parse_survey(wb_path: Path) -> tuple[list[dict], list[ImportIssue]]:
    """Parse the survey workbook.

    Returns a list of parsed survey row dicts (not Equipment objects yet —
    matching to base records happens in the pipeline) and any issues.

    Each dict has:
        asset_number, serial_number, manufacturer, model, description,
        location, blue_dot, how_old, section, source_row
    """
    wb = openpyxl.load_workbook(str(wb_path), data_only=True)
    ws = wb[wb.sheetnames[0]]

    rows: list[dict] = []
    issues: list[ImportIssue] = []
    current_section = "may_2025_cal"  # default until first section label

    for r in range(1, ws.max_row + 1):
        # Read cell values
        vals = []
        for c in range(1, min(15, ws.max_column + 1)):
            v = ws.cell(row=r, column=c).value
            vals.append(clean_value(v) if v is not None else "")

        # Skip completely empty rows
        if not any(vals):
            continue

        # Check for section label
        section = _detect_section(vals)
        if section is not None:
            current_section = section
            continue

        # Skip header rows
        if _is_header_row(vals):
            continue

        # Skip rows before the data starts (title rows etc.)
        if r < 10:
            continue

        # Extract fields by column position
        def col(name: str) -> str:
            idx = SURVEY_COLS.get(name, 0)
            return vals[idx - 1] if idx > 0 and idx <= len(vals) else ""

        asset = col("asset_number")
        serial = col("serial_number")
        mfg = col("manufacturer")
        model = col("model")
        desc = col("description")
        location = col("location")
        blue_dot = col("blue_dot")
        how_old = col("how_old")

        # Skip rows that are just section labels disguised as data
        if not serial and not mfg and not model and not desc:
            if asset and any(kw in asset.lower() for kw in ("cal", "equip", "working", "test")):
                continue

        # Flag placeholder asset numbers
        if asset and asset.lower() in ("need asset no", "need asset no."):
            issues.append(ImportIssue(
                issue_type="placeholder",
                source_file=SURVEY_FILE, source_sheet="Sheet1", source_row=r,
                asset_number=asset, serial_number=serial,
                summary=f"Survey row {r}: placeholder asset number '{asset}'",
                raw_data=json.dumps({
                    "asset_number": asset, "serial_number": serial,
                    "manufacturer": mfg, "model": model, "description": desc,
                }),
            ))

        row_data = {
            "asset_number": asset if not is_placeholder(asset) else "",
            "serial_number": serial if not is_placeholder(serial) else "",
            "manufacturer": mfg,
            "manufacturer_normalized": normalize_manufacturer(mfg),
            "model": model,
            "description": desc,
            "location": location,
            "blue_dot": blue_dot,
            "how_old": how_old.strip() if how_old else "",
            "section": current_section,
            "source_row": r,
        }

        # Only include rows that have at least some identifying info
        if asset or serial or (model and mfg):
            rows.append(row_data)

    return rows, issues


def index_survey_raw_cells(wb_path: Path) -> list[RawCell]:
    """Index every non-empty cell in the survey for raw search."""
    wb = openpyxl.load_workbook(str(wb_path), data_only=True)
    ws = wb[wb.sheetnames[0]]
    cells: list[RawCell] = []

    for r in range(1, ws.max_row + 1):
        row_vals = []
        for c in range(1, min(15, ws.max_column + 1)):
            v = ws.cell(row=r, column=c).value
            cv = clean_value(v) if v is not None else ""
            if cv:
                row_vals.append(cv)

        row_preview = " | ".join(row_vals[:6])

        for c in range(1, min(15, ws.max_column + 1)):
            v = ws.cell(row=r, column=c).value
            cv = clean_value(v) if v is not None else ""
            if cv and not cv.isspace():
                cells.append(RawCell(
                    source_file=SURVEY_FILE,
                    source_sheet="Sheet1",
                    row_number=r,
                    column_number=c,
                    cell_address=f"{col_to_letter(c - 1)}{r}",
                    cell_value=cv,
                    row_preview=row_preview,
                ))

    return cells

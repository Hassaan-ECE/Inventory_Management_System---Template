"""Export managed inventory to a formatted Excel workbook."""

from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.utils import get_column_letter

from app_config import APP_CONFIG
from Code.db.database import get_all_equipment, get_all_import_issues, get_equipment_stats


# ── Shared palette ───────────────────────────────────────────────────────────

_WHITE = "FFFFFF"
_OFF_WHITE = "F9FAFB"
_BAND = "F3F4F6"
_BORDER_COLOR = "D1D5DB"
_HEADER_BG = "1F2937"
_SECTION_BG = "E5E7EB"

_THIN_SIDE = Side(style="thin", color=_BORDER_COLOR)
_CELL_BORDER = Border(left=_THIN_SIDE, right=_THIN_SIDE,
                      top=_THIN_SIDE, bottom=_THIN_SIDE)
_HEADER_BORDER = Border(
    left=Side(style="thin", color="374151"),
    right=Side(style="thin", color="374151"),
    top=Side(style="thin", color="374151"),
    bottom=Side(style="medium", color="374151"),
)

_HEADER_FILL = PatternFill(start_color=_HEADER_BG, end_color=_HEADER_BG, fill_type="solid")
_HEADER_FONT = Font(bold=True, size=11, color=_WHITE)
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

_EVEN_FILL = PatternFill(start_color=_WHITE, end_color=_WHITE, fill_type="solid")
_ODD_FILL = PatternFill(start_color=_BAND, end_color=_BAND, fill_type="solid")

_LEFT = Alignment(horizontal="left", vertical="center")
_CENTER = Alignment(horizontal="center", vertical="center")
_RIGHT = Alignment(horizontal="right", vertical="center")
_WRAP_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

_BODY_FONT = Font(size=10, color="1F2937")
_MUTED_FONT = Font(size=10, color="6B7280")

# ── Status fills (light, print-friendly) ────────────────────────────────────

_LIFECYCLE_FILLS = {
    "active":   PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid"),
    "repair":   PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"),
    "scrapped": PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid"),
    "missing":  PatternFill(start_color="FCE7F3", end_color="FCE7F3", fill_type="solid"),
    "rental":   PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid"),
}

_LIFECYCLE_FONTS = {
    "active":   Font(size=10, color="166534"),
    "repair":   Font(size=10, color="92400E"),
    "scrapped": Font(size=10, color="991B1B"),
    "missing":  Font(size=10, color="9D174D"),
    "rental":   Font(size=10, color="1E40AF"),
}

_WORKING_FILLS = {
    "working":     PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid"),
    "limited":     PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"),
    "not_working": PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid"),
}

_WORKING_FONTS = {
    "working":     Font(size=10, color="166534"),
    "limited":     Font(size=10, color="92400E"),
    "not_working": Font(size=10, color="991B1B"),
}

_CAL_FILLS = {
    "calibrated":     PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid"),
    "out_to_cal":     PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"),
    "reference_only": PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid"),
}

_CAL_FONTS = {
    "calibrated":     Font(size=10, color="166534"),
    "out_to_cal":     Font(size=10, color="92400E"),
    "reference_only": Font(size=10, color="1E40AF"),
}

_ISSUE_FILLS = {
    "duplicate":   PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"),
    "unmatched":   PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid"),
    "missing_id":  PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid"),
    "placeholder": PatternFill(start_color="F3E8FF", end_color="F3E8FF", fill_type="solid"),
    "conflict":    PatternFill(start_color="FCE7F3", end_color="FCE7F3", fill_type="solid"),
    "parse_error": PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid"),
}


# ── Inventory column definitions ─────────────────────────────────────────────
# (header, attribute, width, alignment, special_format)

_INV_COLUMNS = [
    ("Asset Number",    "asset_number",         16, "left",   None),
    ("Serial Number",   "serial_number",        20, "left",   None),
    ("Manufacturer",    "manufacturer",         18, "left",   None),
    ("Model",           "model",                16, "left",   None),
    ("Description",     "description",          32, "left",   None),
    ("Project",         "project_name",         20, "left",   None),
    ("Location",        "location",             24, "left",   None),
    ("Links",           "links",                28, "left",   None),
    ("Assigned To",     "assigned_to",          14, "left",   None),
    ("Lifecycle",       "lifecycle_status",      13, "center", "lifecycle"),
    ("Working",         "working_status",        13, "center", "working"),
    ("Condition",       "condition",             22, "left",   None),
    ("Cal Status",      "calibration_status",    15, "center", "calibration"),
    ("Last Cal Date",   "last_calibration_date", 14, "center", None),
    ("Cal Due Date",    "calibration_due_date",  14, "center", None),
    ("Cal Vendor",      "calibration_vendor",    16, "left",   None),
    ("Cal Cost",        "calibration_cost",      11, "right",  "currency"),
    ("Ownership",       "ownership_type",        12, "center", None),
    ("Rental Vendor",   "rental_vendor",         14, "left",   None),
    ("Rental Cost/Mo",  "rental_cost_monthly",   14, "right",  "currency"),
    ("Verified",        "verified_in_survey",    14, "center", "boolean"),
    ("Blue Dot",        "blue_dot_ref",          10, "center", None),
    ("Est. Age (Yrs)",  "estimated_age_years",   13, "center", "number"),
    ("Notes",           "notes",                 40, "left",   None),
]

_ISSUE_COLUMNS = [
    ("Type",          "issue_type",         14, "center", "issue_type"),
    ("Source File",   "source_file",        16, "left",   None),
    ("Sheet",         "source_sheet",       16, "left",   None),
    ("Row",           "source_row",          8, "center", None),
    ("Asset Number",  "asset_number",       16, "left",   None),
    ("Serial Number", "serial_number",      18, "left",   None),
    ("Summary",       "summary",            50, "left",   None),
    ("Status",        "resolution_status",  13, "center", None),
]


# ── Public API ───────────────────────────────────────────────────────────────

def export_inventory(conn, output_path: Path) -> Path:
    """Export the full inventory to a formatted Excel workbook.

    Creates three sheets:
        Inventory      — all equipment records with status coloring
        Import Issues  — flagged rows from the import pipeline
        Export Summary — record counts and source file metadata

    Returns the output path.
    """
    wb = openpyxl.Workbook()

    _build_inventory_sheet(wb, conn)
    _build_issues_sheet(wb, conn)
    _build_summary_sheet(wb, conn)

    wb.save(str(output_path))
    return output_path


# ── Inventory sheet ──────────────────────────────────────────────────────────

def _build_inventory_sheet(wb: openpyxl.Workbook, conn) -> None:
    ws = wb.active
    ws.title = "Inventory"
    ws.sheet_properties.tabColor = "2563EB"

    _write_header(ws, _INV_COLUMNS)

    equipment = get_all_equipment(conn)
    align_map = {"left": _LEFT, "center": _CENTER, "right": _RIGHT}

    for r, eq in enumerate(equipment, start=2):
        band = _EVEN_FILL if r % 2 == 0 else _ODD_FILL

        for c, (_, attr, _, align_key, fmt) in enumerate(_INV_COLUMNS, start=1):
            raw = getattr(eq, attr, "")
            value, cell_fill, cell_font = _format_cell_value(raw, fmt, band)

            cell = ws.cell(row=r, column=c, value=value)
            cell.font = cell_font or _BODY_FONT
            cell.fill = cell_fill or band
            cell.border = _CELL_BORDER
            cell.alignment = align_map.get(align_key, _LEFT)

            if fmt == "currency" and value is not None:
                cell.number_format = '$#,##0.00'
            elif fmt == "number" and value is not None:
                cell.number_format = '0.0'

    _finalize_sheet(ws, _INV_COLUMNS, len(equipment))


# ── Import Issues sheet ──────────────────────────────────────────────────────

def _build_issues_sheet(wb: openpyxl.Workbook, conn) -> None:
    ws = wb.create_sheet("Import Issues")
    ws.sheet_properties.tabColor = "F59E0B"

    _write_header(ws, _ISSUE_COLUMNS)

    issues = get_all_import_issues(conn)
    align_map = {"left": _LEFT, "center": _CENTER, "right": _RIGHT}

    for r, issue in enumerate(issues, start=2):
        band = _EVEN_FILL if r % 2 == 0 else _ODD_FILL

        for c, (_, attr, _, align_key, fmt) in enumerate(_ISSUE_COLUMNS, start=1):
            raw = getattr(issue, attr, "")
            value, cell_fill, cell_font = _format_cell_value(raw, fmt, band)

            cell = ws.cell(row=r, column=c, value=value)
            cell.font = cell_font or _BODY_FONT
            cell.fill = cell_fill or band
            cell.border = _CELL_BORDER
            cell.alignment = align_map.get(align_key, _LEFT)

    _finalize_sheet(ws, _ISSUE_COLUMNS, len(issues))


# ── Export Summary sheet ─────────────────────────────────────────────────────

def _build_summary_sheet(wb: openpyxl.Workbook, conn) -> None:
    ws = wb.create_sheet("Export Summary")
    ws.sheet_properties.tabColor = "10B981"

    stats = get_equipment_stats(conn)
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    title_font = Font(bold=True, size=16, color="1F2937")
    subtitle_font = Font(size=11, color="6B7280")
    section_font = Font(bold=True, size=12, color="1F2937")
    section_fill = PatternFill(start_color=_SECTION_BG, end_color=_SECTION_BG, fill_type="solid")
    label_font = Font(size=11, color="374151")
    value_font = Font(bold=True, size=11, color="1F2937")

    row = 1

    # Title
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    title_cell = ws.cell(row=row, column=1, value=f"{APP_CONFIG.display_name} — Export Summary")
    title_cell.font = title_font
    title_cell.alignment = Alignment(vertical="center")
    row += 1

    ws.cell(row=row, column=1, value=f"Generated {now}")
    ws.cell(row=row, column=1).font = subtitle_font
    row += 2

    # Inventory Statistics section
    row = _summary_section(ws, row, "Inventory Statistics", section_font, section_fill, [
        ("Total Equipment Records", stats["total"],            value_font),
        ("Active",                  stats["active"],           value_font),
        ("In Repair",               stats["repair"],           value_font),
        ("Scrapped",                stats["scrapped"],         value_font),
        ("Missing",                 stats["missing"],          value_font),
    ], label_font)
    row += 1

    # Calibration section
    row = _summary_section(ws, row, "Calibration", section_font, section_fill, [
        ("Calibrated",              stats["calibrated"],       value_font),
        ("Reference Only (No Cal)", stats["reference_only"],   value_font),
    ], label_font)
    row += 1

    # Audit section
    row = _summary_section(ws, row, "Audit", section_font, section_fill, [
        ("Verified",                stats["verified_in_survey"], value_font),
        ("Unresolved Import Issues", stats["import_issues"],    value_font),
    ], label_font)
    row += 1

    # Source Files section
    row = _summary_section(ws, row, "Source Files", section_font, section_fill, [
        ("Master List", APP_CONFIG.master_source_file, label_font),
        ("Survey",      APP_CONFIG.survey_source_file, label_font),
    ], label_font)

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 52

    # Light borders around each data cell
    for r in range(1, row):
        for c in (1, 2):
            cell = ws.cell(row=r, column=c)
            if cell.value is not None:
                cell.border = Border(
                    bottom=Side(style="hair", color=_BORDER_COLOR),
                )


# ── Helpers ──────────────────────────────────────────────────────────────────

def _write_header(ws, columns: list) -> None:
    """Write a formatted header row with dark fill and white bold text."""
    ws.row_dimensions[1].height = 28

    for c, (header, _, width, _, _) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=c, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
        cell.border = _HEADER_BORDER
        ws.column_dimensions[get_column_letter(c)].width = width


def _finalize_sheet(ws, columns: list, data_row_count: int) -> None:
    """Apply freeze panes, auto-filter, and print setup."""
    ws.freeze_panes = "A2"
    if data_row_count > 0:
        last_col = get_column_letter(len(columns))
        ws.auto_filter.ref = f"A1:{last_col}{data_row_count + 1}"

    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(
        fitToPage=True,
    )
    ws.print_title_rows = "1:1"


def _format_cell_value(raw, fmt: str | None, band_fill: PatternFill):
    """Return (display_value, fill_override, font_override) for a cell.

    Returns None for fill/font when the default band should be used.
    """
    if fmt == "boolean":
        return ("Yes" if raw else "", None, _BODY_FONT if raw else _MUTED_FONT)

    if fmt == "currency":
        if raw is None:
            return (None, None, _MUTED_FONT)
        return (raw, None, None)

    if fmt == "number":
        if raw is None:
            return (None, None, _MUTED_FONT)
        return (raw, None, None)

    if fmt == "lifecycle":
        val = str(raw or "")
        fill = _LIFECYCLE_FILLS.get(val)
        font = _LIFECYCLE_FONTS.get(val, _BODY_FONT)
        return (val, fill, font)

    if fmt == "working":
        val = str(raw or "")
        fill = _WORKING_FILLS.get(val)
        font = _WORKING_FONTS.get(val, _BODY_FONT)
        return (val, fill, font)

    if fmt == "calibration":
        val = str(raw or "")
        fill = _CAL_FILLS.get(val)
        font = _CAL_FONTS.get(val, _BODY_FONT)
        return (val, fill, font)

    if fmt == "issue_type":
        val = str(raw or "")
        fill = _ISSUE_FILLS.get(val)
        return (val, fill, None)

    # Default: plain text
    val = str(raw) if raw else ""
    if not val:
        return ("", None, _MUTED_FONT)
    return (val, None, None)


def _summary_section(ws, start_row: int, title: str, section_font: Font,
                     section_fill: PatternFill, rows: list, label_font: Font) -> int:
    """Write a titled section in the summary sheet. Returns the next available row."""
    row = start_row

    # Section header spanning both columns
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    header_cell = ws.cell(row=row, column=1, value=title)
    header_cell.font = section_font
    header_cell.fill = section_fill
    header_cell.alignment = Alignment(vertical="center")
    ws.cell(row=row, column=2).fill = section_fill
    row += 1

    # Data rows
    for label, value, font in rows:
        label_cell = ws.cell(row=row, column=1, value=label)
        label_cell.font = label_font
        label_cell.alignment = Alignment(indent=1, vertical="center")

        value_cell = ws.cell(row=row, column=2, value=value)
        value_cell.font = font
        row += 1

    return row
